import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Color
from copy import copy

def is_highlighted(cell):
    if cell.fill.fill_type not in [None, "none"] and cell.fill.start_color.index not in ["00000000", "FFFFFF"]:
        return True
    return False

def get_fill_color(count, max_count):
    if count == 1:
        return PatternFill(start_color="F5F55F", end_color="F5F55F", fill_type="solid")  # RGB(245, 245, 95)
    else:
        # Calculate the proportion of the count relative to the max count
        proportion = (count - 1) / (max_count - 1)
        # Linearly scale the green component from 245 (for count=1) to 30 (for max_count)
        green_component = int(245 - (proportion * (245 - 30)))
        green_component = max(min(green_component, 245), 30)  # Ensure the value is within bounds
        green_hex = format(green_component, '02X')
        return PatternFill(start_color=f"E8{green_hex}1E", end_color=f"E8{green_hex}1E", fill_type="solid")

directory = os.path.dirname(os.path.realpath(__file__))
output_path = os.path.join(directory, 'Heat_Mappings.xlsx')
if os.path.exists(output_path):
    os.remove(output_path)

aggregated_counts = {}

# Aggregate highlight counts
for filename in os.listdir(directory):
    if filename.endswith('.xlsx') and filename.lower() != 'heat_mappings.xlsx':
        filepath = os.path.join(directory, filename)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            for cell in row:
                cell_id = (cell.row, cell.column)
                if is_highlighted(cell):
                    aggregated_counts[cell_id] = aggregated_counts.get(cell_id, 0) + 1

max_count = max(aggregated_counts.values(), default=0)

template_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') and f.lower() != 'heat_mappings.xlsx']
if template_files:
    template_path = os.path.join(directory, template_files[0])
    template_wb = openpyxl.load_workbook(template_path)
    template_sheet = template_wb.active

    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = template_sheet.title
    output_sheet.freeze_panes = 'A2'
    top_row_fill = PatternFill(start_color=Color(rgb='E3E1DC'), end_color=Color(rgb='E3E1DC'), fill_type="solid")

    for row in template_sheet.iter_rows():
        for cell in row:
            new_cell = output_sheet.cell(row=cell.row, column=cell.column)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                if cell.row == 1:
                    new_cell.fill = top_row_fill
                else:
                    new_cell.fill = PatternFill(fill_type=None)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            
            cell_id = (cell.row, cell.column)
            if cell_id in aggregated_counts:
                count = aggregated_counts[cell_id]
                new_cell.value = f"{cell.value} ({count})"
                new_cell.fill = get_fill_color(count, max_count)
            else:
                new_cell.value = cell.value

    output_wb.save(output_path)
