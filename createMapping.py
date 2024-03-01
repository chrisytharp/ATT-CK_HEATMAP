import os
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

def is_highlighted(cell):
    """Check if a cell is considered 'highlighted' with any background color other than no fill or white."""
    if cell.fill.fill_type not in [None, "none"] and cell.fill.start_color.index not in ["00000000", "FFFFFF"]:
        return True
    return False

directory = 'C:\\Users\\cytharp\\Desktop\\MITRE'

# Define the output path and check if the file exists. Delete it if it does.
output_path = os.path.join(directory, 'aggregated_counts.xlsx')
if os.path.exists(output_path):
    os.remove(output_path)

aggregated_counts = {}

# Aggregate highlight counts
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        filepath = os.path.join(directory, filename)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            for cell in row:
                cell_id = (cell.row, cell.column)
                if is_highlighted(cell):  # Checking highlight based on background color
                    aggregated_counts[cell_id] = aggregated_counts.get(cell_id, 0) + 1

# Now, copy the formatting and content from a template file and update with counts
template_path = os.path.join(directory, os.listdir(directory)[0])
template_wb = openpyxl.load_workbook(template_path)
template_sheet = template_wb.active

output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
output_sheet.title = template_sheet.title

# Copy cell styles and values from the template, then append counts
for row in template_sheet.iter_rows():
    for cell in row:
        new_cell = output_sheet.cell(row=cell.row, column=cell.column)
        if cell.has_style:  # Copy style if cell has any, excluding fill
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = PatternFill(fill_type=None)  # Reset fill to default (no fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
        
        cell_id = (cell.row, cell.column)
        if cell_id in aggregated_counts:
            # Append the highlight count to the cell's original content
            new_cell.value = f"{cell.value} ({aggregated_counts[cell_id]})"
        else:
            new_cell.value = cell.value

# Save the output workbook
output_wb.save(output_path)
